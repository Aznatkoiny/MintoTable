"""
Excel file creation service
"""

import os
import openpyxl
from datetime import datetime
from models.enums import Stream, Substream, Initiative, ItemType, Stage
from models.project_models import Minutes
from config.app_config import TIMESTAMP_FORMAT, DEFAULT_FILENAME_PREFIX, OUTPUT_DIR

def _get_enum_values(enum_class):
    """Get all values from an enum class"""
    return [e.value for e in enum_class]

def _add_meeting_overview(workbook, minutes_data):
    """
    Add meeting overview sheet to the workbook
    
    Args:
        workbook: Excel workbook
        minutes_data: Minutes object
    """
    main_sheet = workbook.active
    main_sheet.title = "Meeting Overview"
    
    # Add meeting details
    main_sheet['A1'] = "Meeting Title"
    main_sheet['B1'] = minutes_data.meeting_title or "Not specified"
    main_sheet['A2'] = "Date"
    main_sheet['B2'] = minutes_data.meeting_date or "Not specified"
    main_sheet['A3'] = "Summary"
    main_sheet['B3'] = minutes_data.summary or "Not provided"
    
    # Add attendees
    main_sheet['A5'] = "Attendees"
    for i, attendee in enumerate(minutes_data.attendees, start=1):
        main_sheet[f'B{4+i}'] = attendee

def _add_project_items_sheet(workbook, minutes_data):
    """
    Add project items sheet to the workbook
    
    Args:
        workbook: Excel workbook
        minutes_data: Minutes object
    """
    if not minutes_data.items:
        return
    
    items_sheet = workbook.create_sheet("Project Items")
    
    # Define headers (all fields from ProjectItem)
    headers = [
        "TaskID", "Stream", "Substream", "Initiative", "Type", "WorkItem", 
        "Description", "AssignedTo", "Progress", "Priority", "StartDate", 
        "DueDate", "FinishDate", "Stage", "Sprint", "JiraID", "KeyStakeholders", 
        "RAIDTags", "Source", "LinkToSource", "GanttSwimlane", "GanttItem", "Screenshots"
    ]
    
    # Add headers to the sheet
    for col_idx, header in enumerate(headers, start=1):
        items_sheet.cell(row=1, column=col_idx).value = header
    
    # Add data to the sheet
    for row_idx, item in enumerate(minutes_data.items, start=2):
        item_dict = item.model_dump()
        
        for col_idx, header in enumerate(headers, start=1):
            value = item_dict.get(header)
            
            # Handle list fields
            if isinstance(value, list):
                items_sheet.cell(row=row_idx, column=col_idx).value = ", ".join(value) if value else None
            # Handle enum fields
            elif header in ["Stream", "Substream", "Initiative", "Type", "Stage"] and value is not None:
                items_sheet.cell(row=row_idx, column=col_idx).value = value.value if hasattr(value, 'value') else value
            else:
                items_sheet.cell(row=row_idx, column=col_idx).value = value
    
    # Add data validation for enum fields
    _add_data_validation(items_sheet, minutes_data)

def _add_data_validation(sheet, minutes_data):
    """
    Add data validation to the project items sheet
    
    Args:
        sheet: Excel worksheet
        minutes_data: Minutes object
    """
    num_rows = len(minutes_data.items) + 1  # +1 for header row
    
    # Stream validation (column 2)
    stream_values = _get_enum_values(Stream)
    dv = openpyxl.worksheet.datavalidation.DataValidation(
        type="list", formula1=f'"{",".join(stream_values)}"', allow_blank=True
    )
    sheet.add_data_validation(dv)
    dv.add(f'B2:B{num_rows}')
    
    # Substream validation (column 3)
    substream_values = _get_enum_values(Substream)
    dv = openpyxl.worksheet.datavalidation.DataValidation(
        type="list", formula1=f'"{",".join(substream_values)}"', allow_blank=True
    )
    sheet.add_data_validation(dv)
    dv.add(f'C2:C{num_rows}')
    
    # Initiative validation (column 4)
    initiative_values = _get_enum_values(Initiative)
    dv = openpyxl.worksheet.datavalidation.DataValidation(
        type="list", formula1=f'"{",".join(initiative_values)}"', allow_blank=True
    )
    sheet.add_data_validation(dv)
    dv.add(f'D2:D{num_rows}')
    
    # Type validation (column 5)
    type_values = _get_enum_values(ItemType)
    dv = openpyxl.worksheet.datavalidation.DataValidation(
        type="list", formula1=f'"{",".join(type_values)}"', allow_blank=True
    )
    sheet.add_data_validation(dv)
    dv.add(f'E2:E{num_rows}')
    
    # Stage validation (column 14)
    stage_values = _get_enum_values(Stage)
    dv = openpyxl.worksheet.datavalidation.DataValidation(
        type="list", formula1=f'"{",".join(stage_values)}"', allow_blank=True
    )
    sheet.add_data_validation(dv)
    dv.add(f'N2:N{num_rows}')

def _add_raw_minutes_sheet(workbook, minutes_data):
    """
    Add raw minutes text sheet to the workbook
    
    Args:
        workbook: Excel workbook
        minutes_data: Minutes object
    """
    raw_sheet = workbook.create_sheet("Raw Minutes")
    
    # Add the raw minutes text, split by lines to fit in cells if needed
    lines = minutes_data.raw_text.split('\n')
    for i, line in enumerate(lines, start=1):
        raw_sheet[f'A{i}'] = line

def create_excel(minutes_data, output_path=None):
    """
    Create an Excel file from the structured minutes data
    
    Args:
        minutes_data (Minutes): Structured minutes data
        output_path (str, optional): Path to save the Excel file
        
    Returns:
        str: Path to the created Excel file
    """
    if output_path is None:
        # Create a filename based on meeting title and current date/time
        safe_title = "".join(c if c.isalnum() else "_" for c in minutes_data.meeting_title) if minutes_data.meeting_title else "Meeting"
        timestamp = datetime.now().strftime(TIMESTAMP_FORMAT)
        filename = f"{DEFAULT_FILENAME_PREFIX}_{safe_title}_{timestamp}.xlsx"
        
        # Use output directory if specified
        if OUTPUT_DIR:
            output_path = os.path.join(OUTPUT_DIR, filename)
        else:
            output_path = filename
    
    # Create a workbook
    workbook = openpyxl.Workbook()
    
    # Add different sheets to the workbook
    _add_meeting_overview(workbook, minutes_data)
    _add_project_items_sheet(workbook, minutes_data)
    _add_raw_minutes_sheet(workbook, minutes_data)
    
    # Save the workbook
    workbook.save(output_path)
    return output_path